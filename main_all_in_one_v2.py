import os
import tkinter as tk
from tkinter import PhotoImage
from tkinter import ttk, filedialog, messagebox
import webbrowser
# from scrapdata import scrapurl
# from exportdatatoexcelfile import storescrapeddatatoexcel
# from downloadimagesfromexcel import downloadimages
import openpyxl
import requests
from openpyxl import Workbook, load_workbook
import math
from bs4 import BeautifulSoup
import re



# def clean_urls(urls):
#     cleaned_urls = [url for url in urls if url and url.startswith("https://")]
#     return cleaned_urls

def clean_urls(urls):
    cleaned_urls = []
    for url in urls:
        if "http" in url:
            index = url.index("http")
            space_index = url[index:].find(" ")
            newline_index = url[index:].find("\n")
            if space_index == -1 and newline_index == -1:
                cleaned_urls.append(url[index:])
            elif space_index != -1 and newline_index != -1:
                end_index = min(space_index, newline_index) + index
                cleaned_urls.append(url[index:end_index])
            elif space_index != -1:
                cleaned_urls.append(url[index:index + space_index])
            else:
                cleaned_urls.append(url[index:index + newline_index])
    return cleaned_urls


def scrapurl(url, code):
    bad_urls =[]
    main_product_cat = ''
    product_name = ''
    product_description = ''
    product_price = None
    product_rate = None
    product_rate_number = None
    whole_price = ''
    fraction_price = ''
    product_images_urls = []
    response = requests.get(url, headers={'User-Agent': '', 'Accept-Language': 'en-US, en;q=0.5'})
    soup = BeautifulSoup(response.text, 'html.parser')
    try:    
        try:
            main_product_cat = soup.select('#wayfinding-breadcrumbs_feature_div ul li span a')[0].get_text().strip()
        except:
            pass
        try:
            product_name = soup.select('#productTitle')[0].get_text().strip()
        except:
            pass
        try:
            product_rate = soup.select('#acrPopover span a span')[0].get_text().strip()
        except:
            pass
        try:
            product_rate_number = soup.select('#acrCustomerReviewText')[0].get_text().strip()
        except:
            pass
        try:
            try:
                whole_price = soup.select('.a-price-whole')[0].get_text().strip()[:-1]
            except:
                whole_price = ''
            try:
                fraction_price = soup.select('.a-price-fraction')[0].get_text().strip()
            except:
                fraction_price =''
            if whole_price == '' and fraction_price != '' :
                product_price = '.' + fraction_price
            elif whole_price != '' and fraction_price == '' :
                product_price = whole_price
            else:
                product_price = whole_price + "." + fraction_price  # Combine whole and fraction parts
                product_price = product_price.replace(',', '')  # Remove commas
                product_price = float(product_price)  # Convert to float
                product_price = math.ceil(product_price)  # Convert to float
        except:
            pass
        try:
            product_description = ''
            product_desc = soup.select('#feature-bullets ul li span')[:-1]
            for desc in product_desc:
                product_description = product_description + desc.get_text() + ','
        except:
            pass
        try:
            image_urls = soup.select('#altImages ul li span span span span img')
            for image in image_urls[:6]:
                image_url = image['src']
                if image_url.endswith(".jpg"):
                    image_url = re.sub(r'\._(.*?)_\.', r'._AC_SL1500_.', image_url)
                    product_images_urls.append(image_url)

            product_images_urls = set(product_images_urls)
            product_images_urls = list(product_images_urls)
        except:
            pass
    except:
        bad_urls.append(url)
    return code, url, main_product_cat, product_name, product_price, product_rate, product_rate_number, product_description[:-1], str(product_images_urls)

def storescrapeddatatoexcel(urls, download_path, code):
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = 'LOT Num'
    sheet['B1'] = 'urls'
    sheet['C1'] = 'category'
    sheet['D1'] = 'Name'
    sheet['E1'] = 'Price'
    sheet['F1'] = 'Rate'
    sheet['G1'] = 'RateNumber'
    sheet['H1'] = 'Description'
    sheet['I1'] = 'images'
    code = code
    sheet_name = '/scraped_data' + str(code) + '.xlsx'
    for row, url in enumerate(urls, start=2):
        data = scrapurl(url, code)
        print(url)
        sheet.cell(row=row, column=1).value = data[0]
        sheet.cell(row=row, column=2).value = data[1]
        sheet.cell(row=row, column=3).value = data[2]
        sheet.cell(row=row, column=4).value = data[3]
        sheet.cell(row=row, column=5).value = data[4]
        sheet.cell(row=row, column=6).value = data[5]
        sheet.cell(row=row, column=7).value = data[6]
        sheet.cell(row=row, column=8).value = data[7]
        sheet.cell(row=row, column=9).value = data[8]
        code += 1
     
    workbook.save(download_path + sheet_name)

def downloadimages(file_path, start_code):
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    column_to_extract = 'I'
    column_data = []

    for row in worksheet.iter_rows(values_only=True):
        cell_value = row[ord(column_to_extract) - ord('A')]
        links_list = cell_value.split('\n')
        links_list = [link.strip() for link in links_list if link.strip()]
        column_data.append(links_list)

    corrected_links_list = []
    for links_list in column_data[1:]:
        for link in links_list:
            corrected_links = link.replace("[", "").replace("]", "").replace("'", "").split(',')
            corrected_links_list.append(corrected_links)

    for link_group in corrected_links_list:
        counter = 0
        # download_folder = os.path.dirname(file_path) + '/downloaded_images/' + str(start_code)
        download_folder = os.path.dirname(file_path) + '/downloaded_images/' 
        os.makedirs(download_folder, exist_ok=True)
        for link in link_group:
            try:
                response = requests.get(link)
                filename = os.path.join(download_folder, str(start_code) + " " + '(' + str(counter + 1) + ')' + ".jpg")
                with open(filename, 'wb') as file:
                    file.write(response.content)
                print(f"Image downloaded and saved as {filename}")
                counter += 1
            except:
                print("not valid" + link)
        start_code += 1

def browse_path():
    default_path = os.path.join(os.path.expanduser('~'), 'Downloads')
    path = filedialog.askdirectory()
    if path:
        output_label.config(text=f"Selected path: {path}")
    else:
        path = default_path
        output_label.config(text=f"Selected path: {path}")

def runner():
    urls = urls_text.get("1.0", "end-1c").split('\n')
    code = int(code_entry.get())
    
    path = output_label.cget("text").replace("Selected path: ", "")
    if path == '':
        path = os.path.join(os.path.expanduser('~'), 'Downloads')
    urls = clean_urls(urls)
    print(path)
    print(urls)
    try:
        storescrapeddatatoexcel(urls, path, code)
        file_path = path + '/scraped_data' + str(code) + '.xlsx'
        downloadimages(file_path, code)
        message = "Task completed successfully."
        link_message = "Click OK to open the result location: {}".format(os.path.dirname(file_path))
        result = messagebox.showinfo("Success", message, detail=link_message)
        if result == "ok":
            webbrowser.open(os.path.dirname(file_path))
    except Exception as e:
        messagebox.showerror("Task Error", f"An error occurred: {str(e)}")


root = tk.Tk()
root.title("Scraper")  # Title for the window

# Set window icon
# icon = PhotoImage(file="scraper.png")  # Replace "your_icon.png" with your icon file
# root.iconphoto(True, icon)

# Create a style for buttons
style = ttk.Style()
style.configure('TButton', foreground='#0074a0', background='#000000', font=('Arial', 12))
style.map('TButton', background=[('active', '#005d8c')])

# Create a style for labels
style.configure('TLabel', font=('Arial', 12), foreground='#0074a0')  # labels

# Create a style for entry fields
style.configure('TEntry', font=('Arial', 12))

# Create a style for the main frame
style.configure('TFrame', background='#f0f0f0', relief='ridge', borderwidth=2)

# Create and configure the main frame
main_frame = ttk.Frame(root, padding=20)
main_frame.grid(column=0, row=1, sticky=(tk.W, tk.E, tk.N, tk.S))
root.columnconfigure(0, weight=1)
root.rowconfigure(1, weight=1)

# Title and description
title_label = ttk.Label(root, text="ScrapMe", font=('Arial', 16, 'bold'), foreground='#0074a0')
title_label.grid(row=0, column=0, padx=10, pady=10)

# Labels for URLs and code
urls_label = ttk.Label(main_frame, text="URLs:")
urls_label.grid(row=0, column=0, sticky='w')

code_label = ttk.Label(main_frame, text="image code:")
code_label.grid(row=1, column=0, sticky='w')

# Text widget for multiple URLs
urls_text = tk.Text(main_frame, height=10, width=60,relief='flat')
urls_text.grid(row=0, column=1, padx=5, pady=5)

code_entry = ttk.Entry(main_frame, width=60)
code_entry.grid(row=1, column=1, padx=5, pady=5)

# Button to start the process
start_button = ttk.Button(main_frame, text="Start", command=runner)
start_button.grid(row=2, column=1, pady=10)

# Label for displaying the output
output_label = ttk.Label(main_frame, text="",foreground='#000000', font=('Arial', 12))
output_label.grid(row=3, column=0, columnspan=2, pady=10, sticky='w')
# Button to browse for a directory
browse_button = ttk.Button(main_frame, text="Browse", command=browse_path)
browse_button.grid(row=2, column=0, pady=10)


root.mainloop()






