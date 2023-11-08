import os
import tkinter as tk
from tkinter import PhotoImage
from tkinter import ttk, filedialog, messagebox
import webbrowser
# from scrapdata import scrapurl
# from exportdatatoexcelfile import storescrapeddatatoexcel
# from downloadimagesfromexcel import downloadimages
import openpyxl
import requests
from openpyxl import Workbook, load_workbook
import math
from bs4 import BeautifulSoup
import re



# def clean_urls(urls):
#     cleaned_urls = [url for url in urls if url and url.startswith("https://")]
#     return cleaned_urls

def clean_urls(urls):
    cleaned_urls = []
    for url in urls:
        if "https://a.co" in url or "amazon.c" in url:
            index = url.index("http")
            space_index = url[index:].find(" ")
            newline_index = url[index:].find("\n")
            if space_index == -1 and newline_index == -1:
                cleaned_urls.append(url[index:])
            elif space_index != -1 and newline_index != -1:
                end_index = min(space_index, newline_index) + index
                cleaned_urls.append(url[index:end_index])
            elif space_index != -1:
                cleaned_urls.append(url[index:index + space_index])
            else:
                cleaned_urls.append(url[index:index + newline_index])
        else:
            cleaned_urls.append('')
    print(len(cleaned_urls))
    return cleaned_urls

def scrapurl(url, code,choosen_data):
    bad_urls =[]
    # main_product_cat = ''
    category_name = ''
    product_name = ''
    product_cats=[]
    product_description = ''
    product_price = None
    product_rate = None
    product_rate_number = None
    whole_price = ''
    fraction_price = ''
    product_images_urls = []
    response = requests.get(url, headers={'User-Agent': '', 'Accept-Language': 'en-US, en;q=0.5'})
    soup = BeautifulSoup(response.text, 'html.parser')
    data = [code,url]
    try:   

        try:
            image_urls = soup.select('#altImages ul li span span span span img')
            for image in image_urls[:6]:
                image_url = image['src']
                if image_url.endswith(".jpg"):
                    image_url = re.sub(r'\._(.*?)_\.', r'._AC_SL1500_.', image_url)
                    product_images_urls.append(image_url)

            product_images_urls = set(product_images_urls)
            product_images_urls = list(product_images_urls)
        except:
            pass
        data.append(str(product_images_urls))
        if choosen_data['Category']:
            try:
                category_name = soup.select('#wayfinding-breadcrumbs_feature_div ul li span a')[0].get_text().strip()
                if category_name == '' :
                    category_name = soup.select('#nav-subnav a span')[0].get_text().strip()

                # get categories:
                # product_cats = soup.select('#wayfinding-breadcrumbs_feature_div ul li span a')
                # if product_cats != []:
                #     for category in product_cats:
                #         category_name = category_name + ',' + category.get_text().strip()
                # else:
                #     product_cats = soup.select('#nav-subnav a span')
                #     for category in product_cats:
                #         category_name = category_name + ',' + category.get_text().strip()
            except:
                pass
            data.append(category_name)    
        if choosen_data['Pro Name']:
            try:
                product_name = soup.select('#productTitle')[0].get_text().strip()
            except:
                pass
            data.append(product_name) 
        if choosen_data['Price']:
            try:
                try:
                    whole_price = soup.select('.a-price-whole')[0].get_text().strip()[:-1]
                except:
                    whole_price = ''
                try:
                    fraction_price = soup.select('.a-price-fraction')[0].get_text().strip()
                except:
                    fraction_price =''
                if whole_price == '' and fraction_price != '' :
                    product_price = '.' + fraction_price
                elif whole_price != '' and fraction_price == '' :
                    product_price = whole_price
                else:
                    product_price = whole_price + "." + fraction_price  # Combine whole and fraction parts
                    product_price = product_price.replace(',', '')  # Remove commas
                    product_price = float(product_price)  # Convert to float
                    product_price = math.ceil(product_price)  # Convert to float
            except:
                pass
            data.append(product_price) 
        
        if choosen_data['Rate']:
            try:
                product_rate = soup.select('#acrPopover span a span')[0].get_text().strip()
            except:
                pass
            data.append(product_rate) 
        if choosen_data['RateNumber']:
            try:
                product_rate_number = soup.select('#acrCustomerReviewText')[0].get_text().strip()
            except:
                pass
            data.append(product_rate_number) 
        
        if choosen_data['Description']:
            try:
                product_description = ''
                product_desc = soup.select('#feature-bullets ul li span')[:-1]
                for desc in product_desc:
                    product_description = product_description + desc.get_text() + ','
            except:
                pass
            data.append(product_description[:-1]) 
        
    except:
        bad_urls.append(url)
    # return code, url, category_name, product_name, product_price, product_rate, product_rate_number, product_description[:-1], str(product_images_urls)
    return data


from openpyxl.utils import get_column_letter, column_index_from_string

def storescrapeddatatoexcel(urls, download_path, code, choosen_data):
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = 'LOT Num'
    sheet['B1'] = 'urls'
    sheet['C1'] = 'images'

    current_column = 'D'
    column_headers = {}
    for key, value in choosen_data.items():
        if value:
            column_headers[key] = current_column
            sheet[current_column + '1'] = key
            current_column = get_column_letter(column_index_from_string(current_column) + 1)


    code = code
    sheet_name = '/scraped_data' + str(code) + '.xlsx'

    for row, url in enumerate(urls, start=2):
        if url != '':
            data = scrapurl(url, code, choosen_data)
            print(url)
            sheet.cell(row=row, column=1).value = data[0] if len(data) > 0 else ''
            sheet.cell(row=row, column=2).value = data[1] if len(data) > 1 else ''
            sheet.cell(row=row, column=3).value = data[2] if len(data) > 2 else ''
            index = 3
            for key, value in column_headers.items():
                sheet.cell(row=row, column=column_index_from_string(value)).value = data[index] if len(data) > index else ''
                index += 1
            code += 1
        else:
            sheet.cell(row=row, column=1).value = ''
            sheet.cell(row=row, column=2).value = url
            sheet.cell(row=row, column=3).value = '[]'
            sheet.cell(row=row, column=4).value = ''
            sheet.cell(row=row, column=5).value = ''
            sheet.cell(row=row, column=6).value = ''
            sheet.cell(row=row, column=7).value = ''
            sheet.cell(row=row, column=8).value = ''
            sheet.cell(row=row, column=9).value = ''

    workbook.save(download_path + sheet_name)

def downloadimages(file_path, start_code,choosen_data):
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    column_to_extract = 'C'
    column_data = []

    for row in worksheet.iter_rows(values_only=True):
        cell_value = row[ord(column_to_extract) - ord('A')]
        links_list = cell_value.split('\n')
        links_list = [link.strip() for link in links_list if link.strip()]
        column_data.append(links_list)

    corrected_links_list = []
    for links_list in column_data[1:]:
        for link in links_list:
            corrected_links = link.replace("[", "").replace("]", "").replace("'", "").split(',')
            corrected_links_list.append(corrected_links)

    for link_group in corrected_links_list:
        counter = 0
        # download_folder = os.path.dirname(file_path) + '/downloaded_images/' + str(start_code)
        download_folder = os.path.dirname(file_path) + '/downloaded_images/' 
        os.makedirs(download_folder, exist_ok=True)
        for link in link_group:
            try:
                response = requests.get(link)
                filename = os.path.join(download_folder, str(start_code) + " " + '(' + str(counter + 1) + ')' + ".jpg")
                with open(filename, 'wb') as file:
                    file.write(response.content)
                print(f"Image downloaded and saved as {filename}")
                counter += 1
            except:
                print("not valid" + link)
        start_code += 1

def browse_path():
    default_path = os.path.join(os.path.expanduser('~'), 'Downloads')
    path = filedialog.askdirectory()
    if path:
        output_label.config(text=f"Selected path: {path}")
    else:
        path = default_path
        output_label.config(text=f"Selected path: {path}")
def runner():
    code = 0
    urls = urls_text.get("1.0", "end-1c").split('\n')
    try:
        code = int(code_entry.get())
    except:
        pass

    path = output_label.cget("text").replace("Selected path: ", "")
    if path == '':
        path = os.path.join(os.path.expanduser('~'), 'Downloads')
    urls = clean_urls(urls)
    print(urls)
    choosen_data = get_checkbox_values()
    print(choosen_data)
    try:
        storescrapeddatatoexcel(urls, path, code,choosen_data)
        file_path = path + '/scraped_data' + str(code) + '.xlsx'
        if download_checkbox.get():
            downloadimages(file_path, code,choosen_data)
        message = "Task completed successfully."
        link_message = "Click OK to open the result location: {}".format(os.path.dirname(file_path))
        result = messagebox.showinfo("Success", message, detail=link_message)
        if result == "ok":
            webbrowser.open(os.path.dirname(file_path))
    except Exception as e:
        messagebox.showerror("Task Error", f"An error occurred: {str(e)}")

def get_checkbox_values():
    checkbox_values = {}
    for i, checkbox in enumerate(checkboxes):
        checkbox_values[checkbox_labels[i]] = checkbox.get()
    return checkbox_values

# Example usage
def print_checkbox_values():
    checkbox_values = get_checkbox_values()
    for label, value in checkbox_values.items():
        print(f"{label}: {value}")


root = tk.Tk()
root.title("Scraper")  # Title for the window








class EntryWithPlaceholder(ttk.Entry):
    def __init__(self, master=None, placeholder="", **kwargs):
        super().__init__(master, **kwargs)
        self.placeholder = placeholder
        self.default_text = placeholder
        self.default_fg_color = self['foreground']
        
        self.bind("<FocusIn>", self.on_focus_in)
        self.bind("<FocusOut>", self.on_focus_out)
        
        self.show_placeholder()
    
    def on_focus_in(self, event):
        if self.get() == self.default_text:
            self.delete(0, tk.END)
            self['foreground'] = self.default_fg_color
    
    def on_focus_out(self, event):
        if not self.get():
            self.show_placeholder()
    
    def show_placeholder(self):
        self.delete(0, tk.END)
        self.insert(0, self.default_text)
        self['foreground'] = 'gray'












# Set window icon
# icon = PhotoImage(file="scraper.png")  # Replace "your_icon.png" with your icon file
# root.iconphoto(True, icon)

# Create a style for buttons
style = ttk.Style()
style.configure('TButton', foreground='#0074a0', background='#000000', font=('Arial', 12))
style.map('TButton', background=[('active', '#005d8c')])

# Create a style for labels
style.configure('TLabel', font=('Arial', 12), foreground='#0074a0')  # labels

# Create a style for entry fields
style.configure('TEntry', font=('Arial', 12))

# Create a style for the main frame
style.configure('TFrame', background='#f0f0f0', relief='ridge', borderwidth=2)

# Create and configure the main frame
main_frame = ttk.Frame(root, padding=20)
main_frame.grid(column=0, row=1, sticky=(tk.W, tk.E, tk.N, tk.S))
root.columnconfigure(0, weight=1)
root.rowconfigure(1, weight=1)

# Title and description
title_label = ttk.Label(root, text="ScrapMe", font=('Arial', 16, 'bold'), foreground='#0074a0')
title_label.grid(row=0, column=0, padx=10, pady=10)

# Labels for URLs and code
urls_label = ttk.Label(main_frame, text="URLs:")
urls_label.grid(row=0, column=0, sticky='w')

# code_label = ttk.Label(main_frame, text="image code:")
# code_label.grid(row=1, column=0, sticky='w')

# Text widget for multiple URLs
urls_text = tk.Text(main_frame, height=10, width=60, relief='flat')
urls_text.grid(row=0, column=1, padx=5, pady=5)

code_entry = EntryWithPlaceholder(main_frame, placeholder="Enter a code", width=20)
code_entry.grid(row=1, column=1, padx=5, pady=5)

# Checkbox for downloading images
download_checkbox = tk.BooleanVar()
download_checkbox.set(True)  # Set the default value to True

download_checkbox_widget = ttk.Checkbutton(main_frame, text="Download images?", variable=download_checkbox)
download_checkbox_widget.grid(row=2, column=2, padx=5, pady=5)

# Button to start the process
start_button = ttk.Button(main_frame, text="Start", command=runner)
start_button.grid(row=2, column=1, pady=10)

# Label for displaying the output
output_label = ttk.Label(main_frame, text="", foreground='#000000', font=('Arial', 12))
output_label.grid(row=3, column=0, columnspan=2, pady=10, sticky='w')

# ButtonOops! It seems that part of the code got cut off. Could you please provide the missing part?
# Button to browse for a directory
browse_button = ttk.Button(main_frame, text="Browse", command=browse_path)
browse_button.grid(row=2, column=0, pady=10)




# Create checkboxes and labels
checkbox_frame = ttk.Frame(main_frame, borderwidth=0, relief='flat')
checkbox_frame.grid(row=0, column=2, rowspan=1, padx=5)

checkboxes = [] 
checkbox_labels = ["Category","Pro Name", "Price", "Rate", "RateNumber", "Description"]

for i in range(6):
    checkbox_var = tk.BooleanVar()
    checkbox = ttk.Checkbutton(checkbox_frame, variable=checkbox_var)
    checkbox.grid(row=i, column=0, sticky='w')
    checkboxes.append(checkbox_var)
    
    checkbox_label = ttk.Label(checkbox_frame, text=checkbox_labels[i], anchor='w')
    checkbox_label.grid(row=i, column=1, padx=5, sticky='w')

root.mainloop()








