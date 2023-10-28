from openpyxl import Workbook
from scrapdata import scrapurl

def storescrapeddatatoexcel(urls,download_path,code):
    workbook = Workbook()
    sheet = workbook.active
# url, product_name, product_price, product_rate, product_rate_number, str(product_images_urls), product_description[:-1]
    sheet['A1'] = 'LOT Num'
    sheet['B1'] = 'urls'
    sheet['C1'] = 'category'
    sheet['D1'] = 'Name'
    sheet['E1'] = 'Price'
    sheet['F1'] = 'Rate'
    sheet['G1'] = 'RateNumber'
    sheet['H1'] = 'Description'
    sheet['I1'] = 'images'
    code = code
    for row, url in enumerate(urls, start=2):
        
        data = scrapurl(url,code)
        if url != '' :
            sheet.cell(row=row, column=1).value = data[0]
            sheet.cell(row=row, column=2).value = data[1]
            sheet.cell(row=row, column=3).value = data[2]
            sheet.cell(row=row, column=4).value = data[3]
            sheet.cell(row=row, column=5).value = data[4]
            sheet.cell(row=row, column=6).value = data[5]
            sheet.cell(row=row, column=7).value = data[6]
            sheet.cell(row=row, column=8).value = data[7]
            sheet.cell(row=row, column=9).value = data[8]
            code+=1
    workbook.save(download_path + '/scraped_data.xlsx')



