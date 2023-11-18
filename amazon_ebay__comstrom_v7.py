import os
import time
import tkinter as tk
from tkinter import PhotoImage
from tkinter import ttk, filedialog, messagebox
import webbrowser
# from scrapdata import scrapurl
# from exportdatatoexcelfile import storescrapeddatatoexcel
# from downloadimagesfromexcel import downloadimages
import openpyxl
import requests
from openpyxl import Workbook, load_workbook
import math
from bs4 import BeautifulSoup
import re
from openpyxl.utils import get_column_letter, column_index_from_string

def clean_excel_url(url):
    cleaned_url = ''
    if "http" in url :
        index = url.index("http")
        space_index = url[index:].find(" ")
        newline_index = url[index:].find("\n")
        if space_index == -1 and newline_index == -1:
            cleaned_url=url[index:]
        elif space_index != -1 and newline_index != -1:
            end_index = min(space_index, newline_index) + index
            cleaned_url=url[index:end_index]
        elif space_index != -1:
            cleaned_url=url[index:index + space_index]
        else:
            cleaned_url=url[index:index + newline_index]
    return cleaned_url

def read_urls_from_excel(file_path,choosen_site):
    # try:
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    result_list = []
    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        key, value = row
        if value != None:
            if choosen_site == '0' :
                if 'amazon' in value or 'a.c' in value:
                    value_list = []
                    value_list.append(key)
                    if value == None:
                        value_list.append('')
                    else:
                        value = clean_excel_url(value)
                        value_list.append(value)
                        result_list.append(value_list)
            elif choosen_site == '1':
                if 'ebay' in value:
                    value_list = []
                    value_list.append(key)
                    if value == None:
                        value_list.append('')
                    else:
                        value = clean_excel_url(value)
                        value_list.append(value)
                        result_list.append(value_list)
            elif choosen_site == '2' :
                if 'comstrom' in value:
                    value_list = []
                    value_list.append(key)
                    if value == None:
                        value_list.append('')
                    else:
                        value = clean_excel_url(value)
                        value_list.append(value)
                        result_list.append(value_list)
            else:
                value_list = []
                value_list.append(key)
                if value == None:
                    value_list.append('')
                else:
                    value = clean_excel_url(value)
                    value_list.append(value)
                    result_list.append(value_list)
    return result_list
    # except Exception as e:
    #     print(f"An error occurred: {e}")
    #     return None

def amazon_urls(urls):
    cleaned_urls = []
    for url in urls:
        if "https://a.co" in url or "amazon.c" in url:
            index = url.index("http")
            space_index = url[index:].find(" ")
            newline_index = url[index:].find("\n")
            if space_index == -1 and newline_index == -1:
                cleaned_urls.append(url[index:])
            elif space_index != -1 and newline_index != -1:
                end_index = min(space_index, newline_index) + index
                cleaned_urls.append(url[index:end_index])
            elif space_index != -1:
                cleaned_urls.append(url[index:index + space_index])
            else:
                cleaned_urls.append(url[index:index + newline_index])
        elif "ebay" in url or "comstrom" in url :
            pass
        else:
            cleaned_urls.append('')
    if cleaned_urls == ['']:
        return []
    else:
        return cleaned_urls

def ebay_urls(urls):
    cleaned_urls = []
    for url in urls:
        if "ebay" in url:
            index = url.index("http")
            space_index = url[index:].find(" ")
            newline_index = url[index:].find("\n")
            if space_index == -1 and newline_index == -1:
                cleaned_urls.append(url[index:])
            elif space_index != -1 and newline_index != -1:
                end_index = min(space_index, newline_index) + index
                cleaned_urls.append(url[index:end_index])
            elif space_index != -1:
                cleaned_urls.append(url[index:index + space_index])
            else:
                cleaned_urls.append(url[index:index + newline_index])
        elif "amazon" in url or "a.c" in url or "comstrom" in url :
            pass
        else:
            cleaned_urls.append('')
    if cleaned_urls == ['']:
        return []
    else:
        return cleaned_urls

def comstrom_urls(urls):
    cleaned_urls = []
    for url in urls:
        if "comstrom" in url:
            index = url.index("http")
            space_index = url[index:].find(" ")
            newline_index = url[index:].find("\n")
            if space_index == -1 and newline_index == -1:
                cleaned_urls.append(url[index:])
            elif space_index != -1 and newline_index != -1:
                end_index = min(space_index, newline_index) + index
                cleaned_urls.append(url[index:end_index])
            elif space_index != -1:
                cleaned_urls.append(url[index:index + space_index])
            else:
                cleaned_urls.append(url[index:index + newline_index])
        elif "amazon" in url or "a.c" in url or "ebay" in url :
            pass
        else:
            cleaned_urls.append('')
    if cleaned_urls == ['']:
        return []
    else:
        return cleaned_urls

def scrapurl(url, code,choosen_data):
    bad_urls =[]
    # main_product_cat = ''
    category_name = ''
    product_name = ''
    product_cats=[]
    product_description = ''
    product_price = None
    product_rate = None
    product_rate_number = None
    whole_price = ''
    fraction_price = ''
    product_images_urls = []
    response = ''
    try:
        response = requests.get(url, headers={'User-Agent': '', 'Accept-Language': 'en-US, en;q=0.5'})
        soup = BeautifulSoup(response.text, 'html.parser')
        print("qqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqq")
        url = response.url
        print(url)
        print("qqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqq")
    except:
        pass
    data = [code,url]
    if "www.ebay" in url:
        try:   
            try:
                images_counter = 0
                image_urls = soup.select('.ux-image-filmstrip-carousel-item img')
                for image in image_urls:
                    image_url = image['src']
                    if images_counter<=5 and image_url.endswith(".jpg"):
                        image_url = re.sub(r's-l64', r's-l1600', image_url)
                        print(image_url)
                        product_images_urls.append(image_url)
                        images_counter +=1
                product_images_urls = set(product_images_urls)
                product_images_urls = list(product_images_urls)
            except:
                pass
            data.append(str(product_images_urls))
            if choosen_data['Category']:
                try:
                    category_name = soup.select('.viexpsvc nav ul li a span')[0].get_text().strip()
                    # if category_name == '' :
                    #     category_name = soup.select('#nav-subnav a span')[0].get_text().strip()
                except:
                    pass
                data.append(category_name)    
            if choosen_data['Description']:
                try:
                    product_name = soup.select('.x-item-title__mainTitle span')[0].get_text().strip()
                except:
                    pass
                data.append(product_name) 
            if choosen_data['Price']:
                try:
                    product_price = soup.select('.x-price-primary span')[0].get_text().strip()
                except:
                    pass
                data.append(product_price) 
            
            if choosen_data['Rate']:
                try:
                    product_rate = soup.select('#acrPopover span a span')[0].get_text().strip()
                except:
                    pass
                data.append(product_rate) 
            if choosen_data['RateNumber']:
                try:
                    product_rate_number = soup.select('.ux-seller-section__item--seller a span')[1].get_text().strip()
                except:
                    pass
                data.append(product_rate_number) 
            
            if choosen_data['Details']:
                try:
                    product_description = ''
                except:
                    pass
                data.append(product_description) 
            
        except:
            bad_urls.append(url)
    
    if "comstrom" in url:
        try:   
            try:
                images_counter = 0
                image_urls = soup.select('.productView-thumbnails li a')
                for image in image_urls:
                    image_url = image['href']
                    if images_counter<=5 and '.jpg?' in image_url:
                        # image_url = re.sub(r's-l64', r's-l1600', image_url)
                        print(image_url)
                        product_images_urls.append(image_url)
                        images_counter +=1
                product_images_urls = set(product_images_urls)
                product_images_urls = list(product_images_urls)
            except:
                pass
            data.append(str(product_images_urls))
            if choosen_data['Category']:
                try:
                    # category_name = soup.select('.disc li')[2].get_text().strip()[9:]
                    categorypart = soup.select('.disc li')
                    category_name = ''
                    for cat in categorypart:
                        print(cat.get_text())
                        print(type(cat.get_text()))
                        if 'ategory' in cat.get_text():
                            category_name = cat.get_text().split(":")[1].strip()
                        elif 'anufacturer' in cat.get_text():
                            category_name = cat.get_text().split(":")[1].strip()
                        else:
                            category_name = ''

                except:
                    pass
                data.append(category_name)    
            if choosen_data['Description']:
                try:
                    product_name = soup.select('.productView-product h1')[0].get_text().strip()
                except:
                    pass
                data.append(product_name) 
            if choosen_data['Price']:
                try:
                    product_price = soup.select('.productView-price div span')[1].get_text().strip()
                except:
                    pass
                data.append(product_price) 
            
            if choosen_data['Rate']:
                try:
                    # product_rate = soup.select('#acrPopover span a span')[0].get_text().strip()
                    product_rate = ''
                except:
                    pass
                data.append(product_rate) 
            if choosen_data['RateNumber']:
                try:
                    # product_rate_number = soup.select('.ux-seller-section__item--seller a span')[1].get_text().strip()
                    product_rate_number = ''
                except:
                    pass
                data.append(product_rate_number) 
            
            if choosen_data['Details']:
                try:
                    product_description = soup.select('#accordion--description div p')[0].get_text().strip()
                except:
                    pass
                data.append(product_description) 
            
        except:
            bad_urls.append(url)
    
    if "https://a.co" in url or "amazon.c" in url:
        max_retries = 50
        retry_delay = 0.25  # seconds

        for retry in range(max_retries):
            amazon_response = requests.get(url, headers={'User-Agent': '', 'Accept-Language': 'en-US, en;q=0.5'})
            if amazon_response.status_code == 200:
                if amazon_response.text.startswith("<!DOCTYPE html>") == False :
                    # Successful amazon_response
                    break
                else:
                    print(f"Attempt {retry + 1} failed. Retrying in {retry_delay} seconds...")
                    time.sleep(retry_delay)

        amazon_soup = BeautifulSoup(amazon_response.text, 'html.parser')

        try:
            image_urls = amazon_soup.select('#altImages ul li span span span span img')
            for image in image_urls[:6]:
                image_url = image['src']
                if image_url.endswith(".jpg"):
                    image_url = re.sub(r'\._(.*?)_\.', r'._AC_SL1500_.', image_url)
                    product_images_urls.append(image_url)

            product_images_urls = set(product_images_urls)
            product_images_urls = list(product_images_urls)
        except:
            pass
        data.append(str(product_images_urls))
        if choosen_data['Category']:
            try:
                category_name = amazon_soup.select('#wayfinding-breadcrumbs_feature_div ul li span a')[0].get_text().strip()
                if category_name == '' :
                    category_name = amazon_soup.select('#nav-subnav a span')[0].get_text().strip()

                # get categories:
                # product_cats = amazon_soup.select('#wayfinding-breadcrumbs_feature_div ul li span a')
                # if product_cats != []:
                #     for category in product_cats:
                #         category_name = category_name + ',' + category.get_text().strip()
                # else:
                #     product_cats = amazon_soup.select('#nav-subnav a span')
                #     for category in product_cats:
                #         category_name = category_name + ',' + category.get_text().strip()
            except:
                pass
            data.append(category_name)    
        if choosen_data['Description']:
            try:
                product_name = amazon_soup.select('#productTitle')[0].get_text().strip()
            except:
                pass
            data.append(product_name) 
        if choosen_data['Price']:
            try:
                try:
                    whole_price = amazon_soup.select('.a-price-whole')[0].get_text().strip()[:-1]
                except:
                    whole_price = ''
                try:
                    fraction_price = amazon_soup.select('.a-price-fraction')[0].get_text().strip()
                except:
                    fraction_price =''
                if whole_price == '' and fraction_price != '' :
                    product_price = '.' + fraction_price
                elif whole_price != '' and fraction_price == '' :
                    product_price = whole_price
                else:
                    product_price = whole_price + "." + fraction_price  # Combine whole and fraction parts
                    product_price = product_price.replace(',', '')  # Remove commas
                    product_price = float(product_price)  # Convert to float
                    product_price = math.ceil(product_price)  # Convert to float
            except:
                pass
            data.append(product_price) 
        
        if choosen_data['Rate']:
            try:
                product_rate = amazon_soup.select('#acrPopover span a span')[0].get_text().strip()
            except:
                pass
            data.append(product_rate) 
        if choosen_data['RateNumber']:
            try:
                product_rate_number = amazon_soup.select('#acrCustomerReviewText')[0].get_text().strip()
            except:
                pass
            data.append(product_rate_number) 
        
        if choosen_data['Details']:
            try:
                product_description = ''
                product_desc = amazon_soup.select('#feature-bullets ul li span')[:-1]
                for desc in product_desc:
                    product_description = product_description + desc.get_text() + ','
            except:
                pass
            data.append(product_description[:-1]) 
        bad_urls.append(url)
    
    print("testttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttt")
    print(data)
    print("testttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttt")
    # return code, url, category_name, product_name, product_price, product_rate, product_rate_number, product_description[:-1], str(product_images_urls)
    return data

def storescrapeddatatoexcel(urls, download_path, code, choosen_data , excel_urls):
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = 'LOT Num'
    sheet['B1'] = 'urls'
    sheet['C1'] = 'images'

    current_column = 'D'
    column_headers = {}
    for key, value in choosen_data.items():
        if value:
            column_headers[key] = current_column
            sheet[current_column + '1'] = key
            current_column = get_column_letter(column_index_from_string(current_column) + 1)

    code = code
    sheet_name = '/scraped_data' + str(code) + '.xlsx'


    if urls != [] :
        for row, url in enumerate(urls, start=2):
            if url != '':
                data = scrapurl(url, code, choosen_data)
                print("testttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttt")
                print(data)
                print("testttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttt")
                print(url)
                sheet.cell(row=row, column=1).value = data[0] if len(data) > 0 else ''
                sheet.cell(row=row, column=2).value = data[1] if len(data) > 1 else ''
                sheet.cell(row=row, column=3).value = data[2] if len(data) > 2 else ''
                index = 3
                for key, value in column_headers.items():
                    sheet.cell(row=row, column=column_index_from_string(value)).value = data[index] if len(data) > index else ''
                    index += 1
                
            else:
                sheet.cell(row=row, column=1).value = ''
                sheet.cell(row=row, column=2).value = url
                sheet.cell(row=row, column=3).value = '[]'
                sheet.cell(row=row, column=4).value = ''
                sheet.cell(row=row, column=5).value = ''
                sheet.cell(row=row, column=6).value = ''
                sheet.cell(row=row, column=7).value = ''
                sheet.cell(row=row, column=8).value = ''
                sheet.cell(row=row, column=9).value = ''
            code += 1

    if excel_urls :
        print("excel file codes & urls : ")
        print("444444444444444444444444444444444444444444444444444444444")
        print(excel_urls)
        print("444444444444444444444444444444444444444444444444444444444")
        for row, row_data in enumerate(excel_urls, start=len(urls)+2):

            print(row_data)
            print("---------------------------------------------------------")
            code = row_data[0]
            url = row_data[1]
            if url != '':
                print(url)
                data = scrapurl(url, code, choosen_data)
                
                sheet.cell(row=row, column=1).value = data[0] if len(data) > 0 else ''
                sheet.cell(row=row, column=2).value = data[1] if len(data) > 1 else ''
                sheet.cell(row=row, column=3).value = data[2] if len(data) > 2 else ''
                index = 3
                for key, value in column_headers.items():
                    sheet.cell(row=row, column=column_index_from_string(value)).value = data[index] if len(data) > index else ''
                    index += 1
                
            else:
                sheet.cell(row=row, column=1).value = ''
                sheet.cell(row=row, column=2).value = url
                sheet.cell(row=row, column=3).value = '[]'
                sheet.cell(row=row, column=4).value = ''
                sheet.cell(row=row, column=5).value = ''
                sheet.cell(row=row, column=6).value = ''
                sheet.cell(row=row, column=7).value = ''
                sheet.cell(row=row, column=8).value = ''
                sheet.cell(row=row, column=9).value = ''

    workbook.save(download_path + sheet_name)

def downloadimages(file_path, start_code,choosen_data):
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    column_to_extract = 'C'
    column_data = []

    for row in worksheet.iter_rows(values_only=True):
        cell_value = row[ord(column_to_extract) - ord('A')]
        links_list = []
        try:
            links_list = cell_value.split('\n')
        except:
            pass
        links_list = [link.strip() for link in links_list if link.strip()]
        column_data.append(links_list)

    corrected_links_list = []
    for links_list in column_data[1:]:
        for link in links_list:
            corrected_links = link.replace("[", "").replace("]", "").replace("'", "").split(',')
            corrected_links_list.append(corrected_links)

    for link_group in corrected_links_list:
        counter = 0
        # download_folder = os.path.dirname(file_path) + '/downloaded_images/' + str(start_code)
        download_folder = os.path.dirname(file_path) + '/downloaded_images/' 
        os.makedirs(download_folder, exist_ok=True)
        for link in link_group:
            try:
                response = requests.get(link)
                filename = os.path.join(download_folder, str(start_code) + " " + '(' + str(counter + 1) + ')' + ".jpg")
                with open(filename, 'wb') as file:
                    file.write(response.content)
                print(f"Image downloaded and saved as {filename}")
                counter += 1
            except:
                print("not valid" + link)
        start_code += 1

def browse_results_folder_path():
    default_path = os.path.join(os.path.expanduser('~'), 'Downloads')
    path = filedialog.askdirectory()
    if path:
        output_label.config(text=f"{path}")
    else:
        path = default_path
        output_label.config(text=f"{path}")

def browse_excel_file_path():
    excelfilepath = filedialog.askopenfilename()
    print("Selected file:", excelfilepath)
    if excelfilepath:
        output_file_label.config(text=f"{excelfilepath}")
    else:
        excelfilepath = ''

def runner():
    code = 0
    urls = []
    try:
        urls = urls_text.get("1.0", "end-1c").split('\n')
    except:
        pass
    amazon_links = amazon_urls(urls)
    ebay_links = ebay_urls(urls)
    comstrom_links = comstrom_urls(urls)
    all_urls = amazon_links + ebay_links + comstrom_links
    urls = all_urls

    # if amazon_links == [''] and ebay_links == [''] and comstrom_links == [''] :
    #     urls = []

    choosen_site = radio_var.get()
    if choosen_site == '0':
        urls = amazon_links
        print("amazon has been choosen")
    elif choosen_site == '1':
        urls = ebay_links
        print("ebay has been choosen")
    elif choosen_site == '2':
        urls = comstrom_links
        print("comstrom has been choosen")
    print("--------------------------------------------------------------------------------------")
    print("directly appended urls : ")
    print("number of urls = " + str(len(urls)))
    print(urls)
    print("--------------------------------------------------------------------------------------")
    try:
        code = int(code_entry.get())
    except:
        pass

    path = output_label.cget("text").replace("Selected path: ", "")
    excelfilepath = output_file_label.cget("text").replace("Selected path: ", "")
    excel_urls = []
    if excelfilepath != '' :
        excel_urls = read_urls_from_excel(excelfilepath,choosen_site)
    else:
        print('empty file excel')
    if path == '':
        path = os.path.join(os.path.expanduser('~'), 'Downloads')
    
    choosen_data = get_checkbox_values()
    try:
        storescrapeddatatoexcel(urls, path, code,choosen_data,excel_urls)
        file_path = path + '/scraped_data' + str(code) + '.xlsx'
        if download_checkbox.get():
            downloadimages(file_path, code,choosen_data)
        message = "Task completed successfully."
        link_message = "Click OK to open the result location: {}".format(os.path.dirname(file_path))
        result = messagebox.showinfo("Success", message, detail=link_message)
        if result == "ok":
            webbrowser.open(os.path.dirname(file_path))
    except Exception as e:
        messagebox.showerror("Task Error", f"An error occurred: {str(e)}")

def get_checkbox_values():
    checkbox_values = {}
    for i, checkbox in enumerate(checkboxes):
        checkbox_values[checkbox_labels[i]] = checkbox.get()
    return checkbox_values

# Example usage
def print_checkbox_values():
    checkbox_values = get_checkbox_values()
    for label, value in checkbox_values.items():
        print(f"{label}: {value}")

root = tk.Tk()
root.title("Scraper")  # Title for the window


def clear_screen():
    urls_text.delete('1.0', tk.END)  # Clear the URLs text widget
    code_entry.delete(0, tk.END)  # Clear the code entry field

    # Uncheck all the checkboxes
    for checkbox in checkboxes:
        checkbox.set(False)

    # Reset the chosen file or folder
    # output_label.cget("text").set('')  # Assuming you have a StringVar for the chosen file or folder

    # Reset the checked radio button
    radio_var.set('')  # Assuming you have a StringVar for the radio buttons



class EntryWithPlaceholder(ttk.Entry):
    def __init__(self, master=None, placeholder="", **kwargs):
        super().__init__(master, **kwargs)
        self.placeholder = placeholder
        self.default_text = placeholder
        self.default_fg_color = self['foreground']
        
        self.bind("<FocusIn>", self.on_focus_in)
        self.bind("<FocusOut>", self.on_focus_out)
        
        self.show_placeholder()
    
    def on_focus_in(self, event):
        if self.get() == self.default_text:
            self.delete(0, tk.END)
            self['foreground'] = self.default_fg_color
    
    def on_focus_out(self, event):
        if not self.get():
            self.show_placeholder()
    
    def show_placeholder(self):
        self.delete(0, tk.END)
        self.insert(0, self.default_text)
        self['foreground'] = 'gray'


# Set window icon
# icon = PhotoImage(file="scraper.png")  # Replace "your_icon.png" with your icon file
# root.iconphoto(True, icon)

# Create a style for buttons
style = ttk.Style()
style.configure('TButton', foreground='#0074a0', background='#000000', font=('Arial', 12))
style.map('TButton', background=[('active', '#005d8c')])

# Create a style for labels
style.configure('TLabel', font=('Arial', 12), foreground='#0074a0')  # labels

# Create a style for entry fields
style.configure('TEntry', font=('Arial', 12))

# Create a style for the main frame
style.configure('TFrame', background='#f0f0f0', relief='ridge', borderwidth=2)

#----------------------------------------------------------------------------------------------------------------------------------------------

# Title 
title_label = ttk.Label(root, text="ScrapMe", font=('Helvetica', 16, 'bold'), foreground='#0074a0')
title_label.grid(row=0, column=0, padx=10, pady=10)

# Create radio buttons and labels
radio_frame = ttk.Frame(root, borderwidth=0, relief='flat')
radio_frame.grid(row=1, column=0, rowspan=1, padx=5, pady=5)

radio_var = tk.StringVar()
radio_labels = ["Amazon", "Ebay", "Comstrom"]
radio_buttons = []
for i in range(3):
    radio_button = ttk.Radiobutton(radio_frame, text=radio_labels[i], variable=radio_var, value=i)
    radio_button.grid(row=0, column=i, padx=5)
    radio_buttons.append(radio_button)

# radio_buttons[2].configure(state="disabled")
# Create and configure the main frame
main_frame = ttk.Frame(root, padding=20)
main_frame.grid(column=0, row=2)
root.columnconfigure(0, weight=1)
root.rowconfigure(1, weight=1)

# Labels for URLs and code
urls_label = ttk.Label(main_frame, text="URLs:")
urls_label.grid(row=0, column=0, sticky='w')

# code_label = ttk.Label(main_frame, text="image code:")
# code_label.grid(row=1, column=0, sticky='w')

# Text widget for multiple URLs
urls_text = tk.Text(main_frame, height=10, width=60, relief='flat')
urls_text.grid(row=0, column=1, padx=5, pady=5)

code_entry = EntryWithPlaceholder(main_frame, placeholder="Enter a code", width=20)
code_entry.grid(row=1, column=1, padx=5, pady=5)

# Checkbox for downloading images
download_checkbox = tk.BooleanVar()
download_checkbox.set(True)  # Set the default value to True

download_checkbox_widget = ttk.Checkbutton(main_frame, text="Download images?", variable=download_checkbox)
download_checkbox_widget.grid(row=1, column=2, padx=5, pady=5)

# Button to start the process
start_button = ttk.Button(main_frame, text="Start", command=runner)
start_button.grid(row=4, column=1, pady=10)


# Button to clear the screen
clear_button = ttk.Button(main_frame, text="Clear", command=clear_screen)
clear_button.grid(row=4, column=0, pady=10)


# Label for displaying the output
output_label = ttk.Label(main_frame, text="", foreground='#000000', font=('Arial', 12))
output_label.grid(row=2, column=1, columnspan=2, pady=10, sticky='w')
# Label for displaying the output file path
output_file_label = ttk.Label(main_frame, text="", foreground='#000000', font=('Arial', 12))
output_file_label.grid(row=3, column=1, columnspan=2, pady=10, sticky='w')

# ButtonOops! It seems that part of the code got cut off. Could you please provide the missing part?
# Button to browse for a directory
browse_folde_button = ttk.Button(main_frame, text="Browse result path", command=browse_results_folder_path)
browse_folde_button.grid(row=2, column=0, pady=10)

browse_fille_button = ttk.Button(main_frame, text=".xlsx excel file path", command=browse_excel_file_path)
browse_fille_button.grid(row=3, column=0, pady=10)

# Create checkboxes and labels
checkbox_frame = ttk.Frame(main_frame, borderwidth=0, relief='flat')
checkbox_frame.grid(row=0, column=2, rowspan=1, padx=5)

checkboxes = [] 
checkbox_labels = ["Category","Description", "Price", "Rate", "RateNumber", "Details"]

for i in range(6):
    checkbox_var = tk.BooleanVar()
    checkbox = ttk.Checkbutton(checkbox_frame, variable=checkbox_var)
    checkbox.grid(row=i, column=0, sticky='w')
    checkboxes.append(checkbox_var)
    
    checkbox_label = ttk.Label(checkbox_frame, text=checkbox_labels[i], anchor='w')
    checkbox_label.grid(row=i, column=1, padx=5, sticky='w')

root.mainloop()


